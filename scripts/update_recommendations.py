#!/usr/bin/env python3
"""
Convert an Excel recommendations file (.xls or .xlsx) to the app CSV format.

Usage:
  python scripts/update_recommendations.py path/to/input.xls \
      --sheet "Sheet1" \
      --output data/recommendations.csv

Expected output columns (case-insensitive, with French aliases accepted):
  - Theme (alias: Thème)
  - Topic (alias: Sujet)
  - Recommendation (alias: Recommandation)
  - Grade
  - Evidence (alias: Preuves)
  - References (alias: Références)
  - Link (alias: Lien)

Rows missing Recommendation or Evidence are dropped (to match app logic).
"""

import argparse
import os
import sys
import shutil
from datetime import datetime
from typing import Dict

import pandas as pd


EXPECTED = [
    "Theme",
    "Topic",
    "Recommendation",
    "Grade",
    "Evidence",
    "References",
    "Link",
]

ALIASES: Dict[str, str] = {
    # normalize to ASCII, lower, no spaces before matching
    "theme": "Theme",
    "thème": "Theme",
    "topic": "Topic",
    "sujet": "Topic",
    "recommendation": "Recommendation",
    "recommandation": "Recommendation",
    "grade": "Grade",
    "evidence": "Evidence",
    "preuves": "Evidence",
    "references": "References",
    "références": "References",
    "referencess": "References",  # common typo guard
    "link": "Link",
    "lien": "Link",
}


def _simplify(name: str) -> str:
    import unicodedata

    s = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii")
    s = s.strip().lower().replace(" ", "").replace("-", "").replace("_", "")
    return s


def choose_engine(path: str) -> str | None:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        # xlrd<2 supports xls; project pins 1.2.0 in requirements
        return "xlrd"
    if ext in {".xlsx", ".xlsm"}:
        return "openpyxl"
    return None  # let pandas decide


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping: Dict[str, str] = {}
    for c in df.columns:
        key = _simplify(c)
        target = ALIASES.get(key)
        if not target:
            # try direct match against EXPECTED
            for exp in EXPECTED:
                if _simplify(exp) == key:
                    target = exp
                    break
        if target:
            mapping[c] = target

    if not mapping:
        return df  # nothing we can do

    df = df.rename(columns=mapping)

    # Only keep expected columns (add missing as empty)
    for col in EXPECTED:
        if col not in df.columns:
            df[col] = ""

    df = df[EXPECTED]

    # Strip whitespace in string columns
    for col in EXPECTED:
        if pd.api.types.is_string_dtype(df[col]) or df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()

    # Drop rows missing critical fields
    df = df.replace({"nan": "", "None": ""})
    df = df.dropna(subset=["Recommendation", "Evidence"], how="any")
    df = df[(df["Recommendation"].astype(str).str.len() > 0) & (df["Evidence"].astype(str).str.len() > 0)]

    return df


def _extract_topic_with_bold_xlsx(xlsx_path: str, sheet) -> list[str] | None:
    """Best-effort extraction of Topic preserving bold via **markers from .xlsx.

    Uses openpyxl to read rich text runs. If rich data is unavailable, returns None.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        wb = load_workbook(xlsx_path, read_only=False, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) and sheet in wb.sheetnames else (
            wb.worksheets[sheet] if isinstance(sheet, int) and 0 <= sheet < len(wb.worksheets) else wb.worksheets[0]
        )

        # Find header row (first non-empty row containing Topic/aliases)
        header_row = None
        headers = []
        for r in ws.iter_rows(min_row=1, max_row=10):
            values = [str(c.value).strip() if c.value is not None else "" for c in r]
            if any(values):
                headers = values
                header_row = r[0].row
                break

        if not headers:
            return None

        # Map columns
        def _simp(s: str) -> str:
            import unicodedata
            s2 = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode("ascii")
            return s2.strip().lower().replace(" ", "").replace("-", "").replace("_", "")

        topic_idx = None
        for idx, name in enumerate(headers):
            key = _simp(name)
            if key in ("topic", "sujet"):
                topic_idx = idx
                break
        if topic_idx is None:
            return None

        topics = []

        # Helper to build text with ** for bold runs if present
        def cell_text_with_bold(cell) -> str:
            v = cell.value
            try:
                # Some builds of openpyxl expose rich text via .value.rich
                runs = getattr(v, "rich", None)
                if isinstance(runs, list) and runs:
                    parts = []
                    for run in runs:
                        txt = getattr(run, "text", "") or ""
                        font = getattr(run, "font", None)
                        bold = False
                        if font is not None:
                            bold = bool(getattr(font, "b", False))
                        parts.append(f"**{txt}**" if bold and txt else txt)
                    return "".join(parts)
            except Exception:
                pass
            # Fallback plain text
            return "" if v is None else str(v)

        # Iterate data rows after header
        for r in ws.iter_rows(min_row=header_row + 1):
            c = r[topic_idx] if topic_idx < len(r) else None
            if c is None:
                topics.append("")
            else:
                topics.append(cell_text_with_bold(c))

        return topics
    except Exception:
        return None


def main() -> int:
    ap = argparse.ArgumentParser(description="Convert Excel recommendations to CSV for the app.")
    ap.add_argument("input", help="Path to .xls/.xlsx file")
    ap.add_argument("--sheet", help="Sheet name or index (default: first)")
    ap.add_argument("--output", default=os.path.join("data", "recommendations.csv"), help="Output CSV path")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        print(f"Input file not found: {args.input}")
        return 2

    engine = choose_engine(args.input)
    sheet = args.sheet
    try:
        # Allow index for sheet
        sheet_name = None
        if sheet is None:
            sheet_name = 0
        else:
            try:
                sheet_name = int(sheet)
            except ValueError:
                sheet_name = sheet

        df = pd.read_excel(args.input, sheet_name=sheet_name, engine=engine)
    except Exception as e:
        print(f"Failed reading Excel: {e}")
        print("Hint: For .xls install xlrd==1.2.0; for .xlsx install openpyxl.")
        return 3

    if isinstance(df, dict):
        # pandas may return dict of DataFrames when reading multiple sheets
        # pick the first sheet
        df = next(iter(df.values()))

    orig_len = len(df)
    # Best-effort: preserve bold in Topic for .xlsx by inserting ** markers
    if isinstance(sheet_name, (int, str)) and choose_engine(args.input) == "openpyxl":
        topics_marked = _extract_topic_with_bold_xlsx(args.input, sheet_name)
        if topics_marked is not None and len(topics_marked) >= len(df):
            # Align sizes: truncate extras if any due to trailing rows
            df["Topic"] = topics_marked[: len(df)]

    # Heuristic: if no explicit bold markers, bold the lead segment before the first colon
    def _heuristic_bold_topic(s: str) -> str:
        if not isinstance(s, str) or not s:
            return s
        if "**" in s or "<strong>" in s or "<b>" in s:
            return s
        # Use either ':' or ' : '
        idx = s.find(":")
        if idx > 0:
            left = s[:idx].strip()
            right = s[idx:]
            if left:
                return f"**{left}**{right}"
        return s

    if "Topic" in df.columns:
        df["Topic"] = df["Topic"].astype(str).map(_heuristic_bold_topic)
    df = normalize_columns(df)

    # Ensure output dir
    out_dir = os.path.dirname(os.path.abspath(args.output)) or "."
    os.makedirs(out_dir, exist_ok=True)

    # Backup existing
    if os.path.exists(args.output):
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = f"{args.output}.{ts}.bak"
        try:
            shutil.copy2(args.output, backup)
            print(f"Backed up previous CSV to: {backup}")
        except Exception as e:
            print(f"Warning: backup failed: {e}")

    try:
        df.to_csv(args.output, index=False)
    except Exception as e:
        print(f"Failed writing CSV: {e}")
        return 4

    print(
        f"Wrote {len(df)} rows (from {orig_len}) to {args.output}. Columns: {', '.join(df.columns)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
